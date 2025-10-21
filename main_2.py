# main.py
import time
import openpyxl
from driver_manager import get_new_driver_with_retries, fun_check_captcha
from search_manager import perform_all_searches, perform_owner_linkedin_search
from owner_api import fun_send_to_api, parse_owner_names_from_response
from data_saver import is_already_processed, append_row_to_csv
import query_generator as que_gen

INPUT_FILE = "data\\df_san_jose_final 1.xlsx"
OUTPUT_FILE = "output\\san_jose_output.csv"

RESTART_DRIVER_EVERY = 10  # restart driver after this many successful processed restaurants

def main():
    driver = None
    try:
        wb = openpyxl.load_workbook(INPUT_FILE)
        ws = wb.active
    except Exception as e:
        print("❌ Error loading Excel:", e)
        return

    headers = [cell.value for cell in ws[1]]
    try:
        title_idx = headers.index("title")
        address_idx = headers.index("address")
    except ValueError as e:
        print("❌ Required columns not found (title/address).", e)
        return

    total = ws.max_row - 1
    print(f"📊 Rows to process: {total}")


    processed_count = 0  # number of saved restaurants (not rows)
    iteration = 0

    try:
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
            iteration += 1
            title = row[title_idx]
            address = row[address_idx]

            print("\n" + "="*60)
            print(f"➡️ [{row_num}/{total}] Processing: {title} @ {address}")
            print("="*60)

            # 1) If already processed (title+address), skip
            if is_already_processed(title, address, OUTPUT_FILE):
                print("⏭️ Already processed (title+address found) — skipping.")
                continue
            one_time_queries = que_gen.generate_one_time_queries(title, address)
            if not driver:
                print("at start")
                driver = get_new_driver_with_retries(one_time_queries)

            base_results = perform_all_searches(driver, title, address, delay_between=2)

            if base_results:
                # 4) Send owner_info to API using your fun_send_to_api method
                owner_search_result = base_results.get("owner_info", {})
                api_resp = None
                try:
                    api_resp = fun_send_to_api(owner_search_result)
                except Exception as e:
                    print("⚠️ Owner API call error:", e)
                    api_resp = None

                owner_names = parse_owner_names_from_response(api_resp)

                # 5) If owner_names found -> produce one saved row per owner_name with a LinkedIn search
                if owner_names:
                    for owner_name in owner_names:
                        owner_linkedin_result = perform_owner_linkedin_search(driver, owner_name, title,
                                                                              delay_between=2)
                        final_record = {
                            "title": title,
                            "address": address,
                            "owner_info": base_results.get("owner_info"),
                            "social_platforms": base_results.get("social_platforms"),
                            "start_date": base_results.get("start_date"),
                            "franchise_info": base_results.get("franchise_info")
                        }
                        saved = append_row_to_csv(final_record, owner_name, owner_linkedin_result, OUTPUT_FILE)
                        if saved:
                            print(f"✅ Saved row for owner '{owner_name}'")
                        else:
                            print(f"❌ Failed to save row for owner '{owner_name}'")
                    # After saving one or more rows for this restaurant, mark as processed (so next run won't redo)
                    processed_count += 1
                else:
                    # No owner names returned, save a single row with empty owner fields
                    final_record = {
                        "title": title,
                        "address": address,
                        "owner_info": base_results.get("owner_info"),
                        "social_platforms": base_results.get("social_platforms"),
                        "start_date": base_results.get("start_date"),
                        "franchise_info": base_results.get("franchise_info")
                    }
                    saved = append_row_to_csv(final_record, "", None, OUTPUT_FILE)
                    if saved:
                        print("✅ Saved row (no owner names returned)")
                        processed_count += 1
                    else:
                        print("❌ Failed to save row (no owner names)")

                # 6) After each processed restaurant, rotate driver every RESTART_DRIVER_EVERY iterations
                if processed_count and (processed_count % RESTART_DRIVER_EVERY == 0):
                    print(f"♻️ Restarting driver after {processed_count} processed restaurants.")
                    try:
                        driver.quit()
                    except Exception:
                        pass
                    driver = get_new_driver_with_retries(one_time_queries)

                time.sleep(1)

    except KeyboardInterrupt:
        print("⏹️ Interrupted by user")
    except Exception as e:
        print("❌ Unexpected error in main loop:", e)
    finally:
        try:
            driver.quit()
        except Exception:
            pass
        print("🔒 Driver closed. Exiting.")

if __name__ == "__main__":
    main()
